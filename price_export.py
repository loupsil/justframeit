from flask import Blueprint, jsonify, request
import xmlrpc.client
import os
import json
import base64
from io import BytesIO, StringIO
from datetime import datetime
import logging
from dotenv import load_dotenv
import pandas as pd
import openpyxl
import re
from utils import log_route_call

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Create blueprint
price_export_bp = Blueprint('price-export', __name__)

def create_log_capture_handler():
    """Create a string handler to capture logs during processing"""
    log_capture_string = StringIO()
    ch = logging.StreamHandler(log_capture_string)
    ch.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    ch.setFormatter(formatter)
    return ch, log_capture_string

# Odoo Configuration
    """
    Log route calls to Odoo ir.logging model.

    Args:
        models: Odoo models proxy (can be None - function will establish connection if needed)
        uid: User ID (can be None - function will establish connection if needed)
        route_name: Name of the route being called
        payload: The request payload (dict)
        server_logs: Captured server logs (string)
        response_data: The final response data returned by the route (dict)

    Returns:
        int: ID of the created log record, or None if logging failed
    """
    try:
        # If models or uid are missing, try to establish connection
        if not models or not uid:
            logger.info(f"Attempting to establish Odoo connection for logging: {route_name}")
            try:
                uid = get_uid()
                models = get_odoo_models()
                logger.info("Successfully established Odoo connection for logging")
            except Exception as conn_error:
                logger.warning(f"Cannot log to Odoo - failed to establish connection. Route: {route_name}, Error: {str(conn_error)}")
                return None

        logger.info(f"Logging route call to Odoo ir.logging model: {route_name}")

        # Format the log entry in a readable way
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Format payload as readable JSON
        payload_formatted = json.dumps(payload, indent=2, ensure_ascii=False)

        # Format response data as readable JSON
        response_formatted = json.dumps(response_data, indent=2, ensure_ascii=False)

        # Determine log level based on response status
        is_error = response_data.get('status') == 'error'
        log_level = 'error' if is_error else 'info'

        # Set appropriate name based on operation type and status
        log_name = 'Price Export Generation'

        # Create the notes content
        notes_content = f"""🕒 Timestamp: {timestamp}
📍 Route: {route_name}
⚠️  Status: {'ERROR' if is_error else 'SUCCESS'}

📋 PAYLOAD:
{payload_formatted}

📤 RESPONSE:
{response_formatted}

📝 SERVER LOGS:
{server_logs}

{'='*80}
"""

        # Create log record in Odoo ir.logging model
        log_vals = {
            'name': log_name,
            'message': response_formatted,
            'x_studio_notes': notes_content,
            'type': 'server',  # Standard type for server logs
            'level': log_level,  # Error level for failed operations, info for successful
            'path': f'Route: {route_name}',  # Required path field - set to route name
            'func': 'log_route_call',  # Function name where logging occurs
            'line': '0',  # Line number (set to 0 since we don't have actual line number)
            'dbname': ODOO_DB  # Database name from environment
        }

        # Try to create the log record in ir.logging model
        log_id = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'ir.logging', 'create', [log_vals])

        logger.info(f"Successfully created log record with ID: {log_id}")
        return log_id

    except Exception as e:
        logger.error(f"Failed to create log record in Odoo: {str(e)}")
        # Don't raise the exception - logging failure shouldn't break the main flow
        return None

# Odoo Configuration
ODOO_URL = os.getenv('JUSTFRAMEIT_ODOO_URL')
ODOO_DB = os.getenv('JUSTFRAMEIT_ODOO_DB')
ODOO_USERNAME = os.getenv('JUSTFRAMEIT_ODOO_USERNAME')
ODOO_API_KEY = os.getenv('JUSTFRAMEIT_ODOO_API_KEY')

def get_odoo_common():
    """Get Odoo common endpoint"""
    try:
        common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common', allow_none=True)
        return common
    except Exception as e:
        logger.error(f"Failed to connect to Odoo common endpoint: {str(e)}")
        raise

def get_odoo_models():
    """Get Odoo models endpoint"""
    try:
        models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object', allow_none=True)
        return models
    except Exception as e:
        logger.error(f"Failed to connect to Odoo models endpoint: {str(e)}")
        raise

def get_uid():
    """Get Odoo user ID"""
    try:
        common = get_odoo_common()
        uid = common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_API_KEY, {})
        if not uid:
            raise Exception("Authentication failed")
        return uid
    except Exception as e:
        logger.error(f"Failed to authenticate with Odoo: {str(e)}")
        raise

def generate_price_export_excel(models, uid, d3_formula=None):
    """
    Generate Excel file using the exact same logic as the Jupyter notebook.
    Creates a new Excel file with 3 tabs populated with Odoo data.
    Returns the Excel file as bytes, total products, total pricelists, and total duration rules.

    Args:
        models: Odoo models proxy
        uid: Odoo user ID
        d3_formula: Optional formula to set in tab 2, cell D3 (e.g., '=B4', '=B5', etc.)
    """
    try:
        logger.info("Starting Excel price-export generation using exact Jupyter notebook logic")

        import time
        import shutil
        from openpyxl import load_workbook

        # =============================================
        # 📋 STEP 1: Fetch products with price computation = Surface or Circumference
        # =============================================
        logger.info("Fetching products with price computation = Surface or Circumference...")

        products = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'product.product', 'search_read',
            [[['x_studio_price_computation', 'in', ['Surface', 'Circumference']]]],  # Filter by price computation
            {
                'fields': [
                    'name',
                    'id',
                    'x_studio_product_code',
                    'x_studio_location_code',
                    'description_ecommerce',
                    'x_studio_price_computation',
                    'standard_price',
                    'x_studio_associated_service',
                    'x_studio_associated_work_center',
                    'x_studio_associated_cost_per_employee_per_hour'
                ],
                # uncomment for testing with a limit
                #'limit': 100
            }
        )

        total_products = len(products)
        logger.info(f"Fetched {total_products} products")

        # =============================================
        # 📋 STEP 2: Create new Excel file from template
        # =============================================
        # Fetch template from Odoo x_configuration.x_studio_price_export_template field
        logger.info("Fetching Excel template from Odoo x_configuration.x_studio_price_export_template")
        config_ids = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'search', [[]]
        )

        if not config_ids:
            raise Exception("No x_configuration record found")

        config_id = config_ids[0]
        config_data = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'read', [config_id], {
                'fields': ['x_studio_price_export_template']
            }
        )

        if not config_data or not config_data[0].get('x_studio_price_export_template'):
            raise Exception("No template found in x_configuration.x_studio_price_export_template field")

        # Decode the base64 template data
        template_base64 = config_data[0]['x_studio_price_export_template']
        logger.info(f"Raw template base64 data length: {len(template_base64) if template_base64 else 0} characters")

        if not template_base64:
            raise Exception("Template data from Odoo is empty")

        try:
            template_bytes = base64.b64decode(template_base64)
            logger.info(f"Decoded template bytes length: {len(template_bytes)} bytes")
        except Exception as e:
            raise Exception(f"Failed to decode template base64 data from Odoo: {str(e)}")

        if len(template_bytes) < 100:  # Basic sanity check for Excel file size
            raise Exception(f"Template data from Odoo is suspiciously small ({len(template_bytes)} bytes), likely corrupted")

        # Generate timestamp for filenames
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Save template to temporary file
        template_file = f'temp_template_{timestamp}.xlsx'
        with open(template_file, 'wb') as f:
            f.write(template_bytes)
        logger.info(f"Saved template from Odoo to temporary file: {template_file} ({len(template_bytes)} bytes)")

        # Validate the template file can be loaded
        try:
            test_wb = load_workbook(template_file, read_only=True)
            test_wb.close()
            logger.info("Template file validation successful")
        except Exception as e:
            # Clean up the corrupted file
            try:
                os.remove(template_file)
            except:
                pass
            raise Exception(f"Template file from Odoo is corrupted and cannot be loaded: {str(e)}")
        output_file = f'justframeit_pricelist_{timestamp}.xlsx'

        # Copy the template to create a new file
        shutil.copy(template_file, output_file)
        logger.info(f"Created new file: {output_file}")

        # Load the newly created Excel file
        wb = load_workbook(output_file)

        # =============================================
        # 📋 STEP 2.5: Modify D3 formula in TAB 2 if specified
        # =============================================
        if d3_formula is not None:
            logger.info(f"Modifying D3 formula in tab 2 to: {d3_formula}")
            if len(wb.sheetnames) > 1:
                ws2 = wb.worksheets[1]
                ws2['D3'] = d3_formula
                logger.info("D3 formula modified successfully")
            else:
                logger.warning("Could not modify D3 formula - second tab not found")

        # =============================================
        # 📋 STEP 3: Fill TAB 1 with products
        # =============================================
        logger.info("Filling TAB 1 with products data...")
        ws1 = wb.worksheets[0]  # First tab (index 0)

        # Starting row is A7 (row 7)
        start_row = 7
        start_time = time.time()

        # Fill data for each product
        for idx, product in enumerate(products):
            current_row = start_row + idx

            # Helper function to convert list/tuple values to string
            def convert_value(value):
                if isinstance(value, (list, tuple)):
                    # If it's a list/tuple, join elements or take the second element (name)
                    if len(value) > 1:
                        return str(value[1])  # Return the name part
                    return str(value[0]) if value else ''
                return value

            # Fill columns A through J with the product data
            ws1[f'A{current_row}'] = convert_value(product.get('name'))
            ws1[f'B{current_row}'] = convert_value(product.get('id'))
            ws1[f'C{current_row}'] = convert_value(product.get('x_studio_product_code'))
            ws1[f'D{current_row}'] = convert_value(product.get('x_studio_location_code'))
            ws1[f'E{current_row}'] = convert_value(product.get('description_ecommerce'))
            ws1[f'F{current_row}'] = convert_value(product.get('x_studio_price_computation'))
            ws1[f'G{current_row}'] = convert_value(product.get('standard_price'))
            ws1[f'H{current_row}'] = convert_value(product.get('x_studio_associated_service'))
            ws1[f'I{current_row}'] = convert_value(product.get('x_studio_associated_work_center'))
            ws1[f'J{current_row}'] = convert_value(product.get('x_studio_associated_cost_per_employee_per_hour'))

            if idx % 50 == 0:  # Log progress every 50 products
                logger.info(f"Filled row {current_row} on tab 1 with product: {product.get('name')}")

        # Save the workbook
        wb.save(output_file)
        logger.info(f"Successfully filled {len(products)} products into tab 1")

        # =============================================
        # 📋 STEP 4: Fetch pricelists and fill TAB 2
        # =============================================
        logger.info("Fetching pricelists and filling TAB 2...")

        # Fetch all pricelists and filter out "Default" in Python to avoid domain parsing issues
        all_pricelists = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'product.pricelist', 'search_read',
            [],  # Empty domain = get all pricelists
            {
                'fields': ['name', 'x_studio_price_discount'],
                'limit': 100  # Add limit to prevent excessive results
            }
        )

        # Filter out pricelists named "Default" (case insensitive)
        pricelists = [
            p for p in all_pricelists
            if p.get('name', '').lower() != 'default'
        ]

        total_pricelists = len(pricelists)
        logger.info(f"Fetched {total_pricelists} pricelists")

        # Access the second tab (sheet)
        if len(wb.sheetnames) > 1:
            ws2 = wb.worksheets[1]
        else:
            logger.warning("Second tab not found in workbook")
            ws2 = None

        if ws2:
            # Starting row is A3 (row 3)
            pricelist_start_row = 3

            # Fill data for each pricelist
            for idx, pricelist in enumerate(pricelists):
                current_row = pricelist_start_row + idx

                # Helper function to convert list/tuple values to string
                def convert_value(value):
                    if isinstance(value, (list, tuple)):
                        if len(value) > 1:
                            return str(value[1])
                        return str(value[0]) if value else ''
                    return value

                # Fill columns A and B with pricelist data
                ws2[f'A{current_row}'] = convert_value(pricelist.get('name'))
                ws2[f'B{current_row}'] = convert_value(pricelist.get('x_studio_price_discount'))

                if idx % 50 == 0:  # Log progress every 50 pricelists
                    logger.info(f"Filled row {current_row} on tab 2 with pricelist: {pricelist.get('name')}")

            # Save the workbook again with pricelist data
            wb.save(output_file)
            logger.info(f"Successfully filled {len(pricelists)} pricelists into tab 2")

        # =============================================
        # 📋 STEP 5: Fetch service duration rules and fill TAB 3
        # =============================================
        logger.info("Fetching service duration rules and filling TAB 3...")

        duration_rules = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'x_services_duration_rules', 'search_read',
            [[]],  # Empty domain = all duration rules
            {
                'fields': [
                    'x_associated_service',
                    'x_studio_work_center',
                    'x_studio_quantity',
                    'x_duurtijd_totaal'
                ]
            }
        )

        total_duration_rules = len(duration_rules)
        logger.info(f"Fetched {total_duration_rules} service duration rules")

        # Access the third tab (sheet)
        if len(wb.sheetnames) > 2:
            ws3 = wb.worksheets[2]
        else:
            logger.warning("Third tab not found in workbook")
            ws3 = None

        if ws3:
            # Starting row is A3 (row 3)
            duration_start_row = 3

            # Fill data for each duration rule
            for idx, rule in enumerate(duration_rules):
                current_row = duration_start_row + idx

                # Helper function to convert list/tuple values to string
                def convert_value(value):
                    if isinstance(value, (list, tuple)):
                        if len(value) > 1:
                            return str(value[1])
                        return str(value[0]) if value else ''
                    return value

                # Fill columns A through D with duration rule data
                ws3[f'A{current_row}'] = convert_value(rule.get('x_associated_service'))
                ws3[f'B{current_row}'] = convert_value(rule.get('x_studio_work_center'))
                ws3[f'C{current_row}'] = convert_value(rule.get('x_studio_quantity'))
                ws3[f'D{current_row}'] = convert_value(rule.get('x_duurtijd_totaal'))

                if idx % 50 == 0:  # Log progress every 50 rules
                    logger.info(f"Filled row {current_row} on tab 3 with service: {rule.get('x_associated_service')}")

            # Save the workbook again with duration rules data
            wb.save(output_file)
            logger.info(f"Successfully filled {len(duration_rules)} service duration rules into tab 3")

        # =============================================
        # 📋 STEP 6: Read the final Excel file and return bytes
        # =============================================
        logger.info("Reading final Excel file to return as bytes...")

        # Read the file back as bytes
        with open(output_file, 'rb') as f:
            excel_bytes = f.read()

        # Clean up temporary files
        try:
            os.remove(template_file)
            logger.info(f"Cleaned up temporary template file: {template_file}")
        except Exception as e:
            logger.warning(f"Failed to clean up temporary template file {template_file}: {str(e)}")

        try:
            os.remove(output_file)
            logger.info(f"Cleaned up temporary output file: {output_file}")
        except Exception as e:
            logger.warning(f"Failed to clean up temporary output file {output_file}: {str(e)}")

        logger.info("Excel file generated successfully with all tabs populated")

        # Return: excel_bytes, products_count, pricelists_count, duration_rules_count, source_description
        return excel_bytes, total_products, total_pricelists, f"Generated from Odoo data: {total_products} products, {total_pricelists} pricelists, {total_duration_rules} duration rules"

    except Exception as e:
        logger.error(f"Error generating price-export Excel: {str(e)}")
        raise

def generate_csv_from_excel(excel_bytes):
    """
    Generate CSV from Excel bytes using the exact same logic as the Jupyter notebook.
    Uses xlwings to force calculations before extracting data.
    Takes Excel bytes, processes the first worksheet, and returns CSV bytes.
    """
    try:
        logger.info("Starting CSV generation from Excel bytes using xlwings calculation")

        import xlwings as xw
        import tempfile

        # Save Excel bytes to temporary file for xlwings
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_excel:
            temp_excel.write(excel_bytes)
            temp_excel_path = temp_excel.name

        logger.info(f"Saved Excel bytes to temporary file: {temp_excel_path}")

        try:
            # Open with xlwings to force calculation
            logger.info("Opening Excel file with xlwings to force calculations")
            app = xw.App(visible=False)
            wb_xlwings = app.books.open(temp_excel_path)
            ws_xlwings = wb_xlwings.sheets[0]

            # Force calculation
            logger.info("Forcing Excel calculation...")
            wb_xlwings.app.calculate()

            # Save the calculated workbook to another temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_calculated:
                temp_calculated_path = temp_calculated.name

            wb_xlwings.save(temp_calculated_path)
            wb_xlwings.close()
            app.quit()

            logger.info(f"Calculated workbook saved as: {temp_calculated_path}")

            # Now load with openpyxl
            logger.info("Loading calculated Excel file with openpyxl")
            wb = openpyxl.load_workbook(temp_calculated_path, data_only=True)
            logger.info("Calculated Excel file loaded successfully")

        finally:
            # Clean up the first temporary file
            try:
                os.remove(temp_excel_path)
                logger.info(f"Cleaned up temporary input file: {temp_excel_path}")
            except Exception as e:
                logger.warning(f"Could not remove temporary input file: {e}")

        # Access the first tab
        ws = wb.worksheets[0]
        logger.info(f"Accessing first worksheet: {ws.title}")

        # Headers are in row 6
        header_row = 6
        start_row = 7  # Data starts from row 7

        # Find the last row with data in column A (starting from row 7)
        last_row = start_row

        logger.info("Finding last row with data...")
        # Find where column A becomes 0 or empty
        for row in range(start_row, ws.max_row + 1):
            if row % 100 == 0:  # Log progress every 100 rows
                logger.info(f"   Checking row {row}...")
            cell_value = ws[f'A{row}'].value
            if cell_value == 0 or cell_value is None or cell_value == '':
                break
            last_row = row

        logger.info(f"Extracting data from row {start_row} to row {last_row}")

        # Get headers from row 6, but check row 5 for dimension indicators
        logger.info(f"Reading headers from row {header_row}...")
        headers = []
        valid_columns = []  # Track which columns to include
        dimension_columns = []  # Track which columns are dimensions
        found_first_dimension = False
        last_dimension_col = 0
        for col in range(1, ws.max_column + 1):
            if col % 50 == 0:  # Log progress every 50 columns
                logger.info(f"   Processing column {col}...")
            # Check row 5 for dimension indicator (e.g., "10.0 x 20.0")
            dimension_cell = ws.cell(row=5, column=col)
            dimension_value = dimension_cell.value

            # Get the header from row 6
            header_cell = ws.cell(row=header_row, column=col)
            header_value = header_cell.value

            # Stricter check: dimension must contain 'x' AND at least one number
            is_dimension = False
            if dimension_value and isinstance(dimension_value, str):
                # Check if it contains 'x' and at least one digit
                if 'x' in dimension_value.lower() and re.search(r'\d', dimension_value):
                    is_dimension = True
                    found_first_dimension = True
                    last_dimension_col = col

            # If we found dimensions before but now stopped finding them, stop processing columns
            if found_first_dimension and not is_dimension:
                logger.info(f"Stopping at column {col} - no more dimension headers found")
                break

            # If row 5 contains a valid dimension indicator, use it
            if is_dimension:
                headers.append(dimension_value)
                valid_columns.append(col)
                dimension_columns.append(len(headers) - 1)  # Track the index of this dimension column
            elif header_value and not found_first_dimension:
                # Only include non-dimension columns if we haven't found dimensions yet
                headers.append(header_value)
                valid_columns.append(col)
            elif not found_first_dimension:
                headers.append(f'Column_{col}')
                valid_columns.append(col)

        logger.info(f"Found {len(headers)} columns (last dimension column: {last_dimension_col})")

        # Extract data from row 7 onwards to the last valid row
        logger.info(f"Extracting data from {last_row - start_row + 1} rows...")
        data = []
        for row in range(start_row, last_row + 1):
            if (row - start_row) % 50 == 0:  # Log progress every 50 rows
                logger.info(f"   Processing row {row} ({row - start_row + 1}/{last_row - start_row + 1})...")
            row_data = []
            for idx, col in enumerate(valid_columns):  # Only extract data from valid columns
                cell = ws.cell(row=row, column=col)
                # Get the actual value, not the formula
                cell_value = cell.value

                # Check if this is a dimension column
                is_dimension_col = idx in dimension_columns

                # Handle special cases
                if cell_value is None:
                    row_data.append(None)
                elif isinstance(cell_value, str) and cell_value.startswith('!REF'):
                    # Handle reference errors
                    row_data.append(None)
                elif isinstance(cell_value, (int, float)):
                    # Round all numeric values to 2 decimal places
                    row_data.append(round(float(cell_value), 2))
                else:
                    row_data.append(cell_value)
            data.append(row_data)

        # Create DataFrame
        logger.info("Creating DataFrame...")
        df = pd.DataFrame(data, columns=headers)
        logger.info(f"DataFrame created with {len(df)} rows and {len(df.columns)} columns")

        # Convert DataFrame to CSV bytes
        logger.info("Converting DataFrame to CSV bytes...")
        csv_buffer = BytesIO()
        df.to_csv(csv_buffer, index=False)
        csv_bytes = csv_buffer.getvalue()

        logger.info(f"Successfully generated CSV with {len(df)} rows from Excel data")

        # Clean up the calculated temporary file
        try:
            os.remove(temp_calculated_path)
            logger.info(f"Cleaned up temporary calculated file: {temp_calculated_path}")
        except Exception as e:
            logger.warning(f"Could not remove temporary calculated file: {e}")

        return csv_bytes

    except Exception as e:
        logger.error(f"Error generating CSV from Excel: {str(e)}")
        raise


def process_pricelist_for_parallel(pricelist_name, pricelist_row, timestamp):
    """
    Process a single pricelist to generate its Excel file.
    This function must be at module level to be pickleable for multiprocessing.
    Establishes its own Odoo connection since connections cannot be pickled.
    Returns Excel bytes - CSV conversion happens in main process due to xlwings limitations.
    """
    try:
        logger.info(f"Processing pricelist '{pricelist_name}' at row {pricelist_row}")

        # Calculate the formula for D3 (pointing to column B of this pricelist row)
        d3_formula = f"=B{pricelist_row}"

        # Establish new Odoo connection for this worker process
        logger.info(f"Establishing Odoo connection for pricelist '{pricelist_name}'")
        worker_uid = get_uid()
        worker_models = get_odoo_models()
        logger.info(f"Successfully connected to Odoo for pricelist '{pricelist_name}'")

        # Add retry logic for Excel generation to handle transient errors
        max_retries = 3
        for attempt in range(max_retries):
            try:
                logger.info(f"Generating Excel for pricelist '{pricelist_name}' (attempt {attempt + 1}/{max_retries})")
                # Generate Excel with modified D3 formula
                modified_excel_bytes, _, _, _ = generate_price_export_excel(worker_models, worker_uid, d3_formula=d3_formula)
                logger.info(f"Successfully generated Excel for pricelist '{pricelist_name}' ({len(modified_excel_bytes)} bytes)")
                break  # Success, exit retry loop
            except Exception as e:
                logger.warning(f"Attempt {attempt + 1} failed for pricelist '{pricelist_name}': {str(e)}")
                if attempt == max_retries - 1:  # Last attempt
                    raise
                # Wait before retry
                import time
                time.sleep(1)

        # Return Excel bytes - CSV conversion will happen in main process
        logger.info(f"Successfully generated Excel for pricelist '{pricelist_name}' - CSV conversion will happen in main process")

        return (pricelist_name, modified_excel_bytes, timestamp)

    except Exception as e:
        logger.error(f"Error processing pricelist '{pricelist_name}': {str(e)}")
        logger.error(f"Exception type: {type(e).__name__}")
        import traceback
        logger.error(f"Exception traceback: {traceback.format_exc()}")
        raise

def generate_csvs_from_pricelists(models, uid, original_excel_bytes, timestamp):
    """
    Generate CSV files by modifying D3 formula to point to different pricelists.
    All pricelists are processed using the same logic with parallelization.

    This function:
    1. Loads the original Excel and reads pricelist names from tab 2, column A starting from row 3
    2. For each pricelist, modifies D3 to point to that row's B column
    3. Generates CSV from each modified Excel using parallel processing
    4. Returns list of (pricelist_name, csv_bytes, csv_filename) tuples

    Args:
        models: Odoo models proxy
        uid: Odoo user ID
        original_excel_bytes: The original Excel file bytes
        timestamp: Timestamp string for filename generation

    Returns:
        list: List of tuples (pricelist_name, csv_bytes, csv_filename)
    """
    try:
        logger.info("Starting generation of CSVs from pricelists")

        import tempfile

        # Save original Excel bytes to temporary file for reading pricelist names
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_excel:
            temp_excel.write(original_excel_bytes)
            temp_excel_path = temp_excel.name

        try:
            # Load Excel to read pricelist names from tab 2
            logger.info("Loading Excel to read pricelist names from tab 2")
            wb = openpyxl.load_workbook(temp_excel_path, data_only=True)

            if len(wb.sheetnames) < 2:
                logger.warning("Second tab not found in workbook, cannot generate CSVs")
                return []

            ws2 = wb.worksheets[1]  # Second tab (pricelists)

            # Read pricelist names from column A starting from row 3
            pricelist_names = []
            row = 3  # Start from row 3

            logger.info("Reading pricelist names from tab 2, column A starting from row 3")

            while row <= ws2.max_row:
                cell_value = ws2[f'A{row}'].value
                if cell_value is None or cell_value == '' or cell_value == 0:
                    break  # Stop when we hit empty cells

                pricelist_name = str(cell_value).strip()
                logger.info(f"Found pricelist at row {row}: '{pricelist_name}'")

                pricelist_names.append((pricelist_name, row))
                logger.info(f"Added pricelist: '{pricelist_name}' at row {row}")

                row += 1

                # Limit to 5 pricelists to prevent excessive processing
                if len(pricelist_names) >= 5:
                    logger.info("Reached limit of 5 pricelists")
                    break

            wb.close()

            logger.info(f"Found {len(pricelist_names)} pricelists to process")

            # Generate CSV for each pricelist using parallel processing
            from concurrent.futures import ProcessPoolExecutor

            csvs = []
            max_workers = min(len(pricelist_names), 5)  # Limit to 5 concurrent Excel processes

            logger.info(f"Starting parallel processing of {len(pricelist_names)} pricelists with {max_workers} workers")

            with ProcessPoolExecutor(max_workers=max_workers) as executor:
                # Submit all pricelist processing tasks and maintain order
                futures = [
                    executor.submit(process_pricelist_for_parallel, pricelist_name, pricelist_row, timestamp)
                    for pricelist_name, pricelist_row in pricelist_names
                ]

                logger.info(f"Submitted {len(futures)} pricelist Excel generation tasks to executor")

                # Collect Excel results from parallel processing
                excel_results = []
                completed_count = 0
                for future in futures:
                    try:
                        result = future.result()
                        completed_count += 1
                        pricelist_name, excel_bytes, result_timestamp = result
                        logger.info(f"Completed Excel generation for pricelist '{pricelist_name}' ({completed_count}/{len(futures)})")
                        excel_results.append((pricelist_name, excel_bytes, result_timestamp))
                    except Exception as exc:
                        # Get the pricelist name from the failed future (need to find which one it was)
                        pricelist_name = "unknown"
                        for i, f in enumerate(futures):
                            if f == future:
                                pricelist_name = pricelist_names[i][0]
                                break
                        logger.error(f'Pricelist {pricelist_name} generated an exception: {str(exc)}')
                        logger.error(f'Exception type: {type(exc).__name__}')
                        import traceback
                        logger.error(f'Exception traceback: {traceback.format_exc()}')
                        raise  # Re-raise to fail the entire operation

            logger.info(f"Parallel Excel generation completed - processed {completed_count} tasks")

            # Now convert Excel files to CSV in main process (xlwings can only work in main process)
            logger.info("Converting Excel files to CSV in main process...")
            for pricelist_name, excel_bytes, result_timestamp in excel_results:
                try:
                    logger.info(f"Converting Excel to CSV for pricelist '{pricelist_name}'")
                    csv_bytes = generate_csv_from_excel(excel_bytes)
                    csv_filename = f"justframeit_price_export_pricelist_{pricelist_name.replace(' ', '_').lower()}_{result_timestamp}.csv"
                    logger.info(f"Successfully converted Excel to CSV for pricelist '{pricelist_name}'")
                    csvs.append((pricelist_name, csv_bytes, csv_filename))
                except Exception as e:
                    logger.error(f"Error converting Excel to CSV for pricelist '{pricelist_name}': {str(e)}")
                    raise

            logger.info(f"CSV conversion completed - generated {len(csvs)} CSV files")
            return csvs

        finally:
            # Clean up temporary file
            try:
                os.remove(temp_excel_path)
                logger.info(f"Cleaned up temporary Excel file: {temp_excel_path}")
            except Exception as e:
                logger.warning(f"Could not remove temporary Excel file: {e}")

    except Exception as e:
        logger.error(f"Error generating CSVs from pricelists: {str(e)}")
        logger.error(f"Exception type: {type(e).__name__}")
        import traceback
        logger.error(f"Exception traceback: {traceback.format_exc()}")
        raise

@price_export_bp.route('/generate-price-export', methods=['POST'])
def generate_price_export():
    """
    Generate Excel price-export and CSV, then save both to Odoo x_configuration fields.

    This route:
    1. Fetches Excel template from x_configuration record ID 1, field x_studio_price_export_template
    2. Processes the template using the exact code from the Jupyter notebook
    3. Saves the generated Excel file to Odoo's x_configuration.x_studio_price_list_1 binary field
    4. Generates CSV from the Excel data using the same logic as the Jupyter notebook (if x_studio_is_run_locally is true)
    5. Saves the generated CSV file to Odoo's x_configuration.x_studio_price_list_1_csv binary field
    6. Returns success/failure status

    POST request with payload containing x_studio_is_run_locally flag:
    POST /generate-price-export
    Content-Type: application/json
    {
        "_action": "CUSTOM - Product Price Export Trigger(#969)",
        "_id": 1,
        "_model": "x_configuration",
        "id": 1,
        "x_studio_is_run_locally": false
    }
    """
    try:
        # Get the request payload
        payload = request.get_json() or {}
        x_studio_is_run_locally = payload.get('x_studio_is_run_locally', True)  # Default to True if not specified

        # Set up log capture
        log_handler, log_capture_string = create_log_capture_handler()
        logger.addHandler(log_handler)

        logger.info("Starting price-export generation route")
        logger.info(f"Payload received: {json.dumps(payload, indent=2)}")
        logger.info(f"x_studio_is_run_locally: {x_studio_is_run_locally}")

        if not x_studio_is_run_locally:
            logger.info("x_studio_is_run_locally is false - CSV generation will be skipped")
        else:
            logger.info("x_studio_is_run_locally is true - CSV generation will proceed")

        # Get Odoo connection
        logger.info("Connecting to Odoo")
        uid = get_uid()
        models = get_odoo_models()
        logger.info("Connected to Odoo successfully")

        # Generate timestamp for filenames
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Generate the Excel file
        logger.info("Generating Excel price-export file")
        excel_bytes, total_products, total_pricelists, source_file = generate_price_export_excel(models, uid)

        # Generate all pricelist-based CSVs using uniform logic (only if x_studio_is_run_locally is true)
        if x_studio_is_run_locally:
            logger.info("Generating pricelist-based CSV files")
            all_pricelist_csvs = generate_csvs_from_pricelists(models, uid, excel_bytes, timestamp)
            logger.info(f"Generated {len(all_pricelist_csvs)} pricelist-based CSV files")
        else:
            logger.info("Skipping CSV generation as x_studio_is_run_locally is false")
            all_pricelist_csvs = []

        # All CSVs follow the same process - no distinction between main and additional
        if all_pricelist_csvs:
            # Use the first CSV as the primary one for backward compatibility
            csv_bytes, csv_filename = all_pricelist_csvs[0][1], all_pricelist_csvs[0][2]
            # All CSVs are treated equally for storage
            additional_csvs = all_pricelist_csvs
        else:
            # Handle case where CSV generation was skipped or no pricelists found
            if x_studio_is_run_locally:
                # Fallback: if no pricelists found, generate a basic CSV (though this shouldn't happen)
                logger.warning("No pricelist CSVs generated, falling back to basic CSV generation")
                csv_bytes = generate_csv_from_excel(excel_bytes)
                csv_filename = f"justframeit_price_export_fallback_{timestamp}.csv"
            else:
                # CSV generation was intentionally skipped
                logger.info("CSV generation was skipped as per x_studio_is_run_locally flag")
                csv_bytes = None
                csv_filename = None
            additional_csvs = []

        # Find the x_configuration record to update
        logger.info("Finding x_configuration record")
        config_ids = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'search', [[]])

        if not config_ids:
            raise Exception("No x_configuration record found")

        config_id = config_ids[0]  # Use the first configuration record
        logger.info(f"Found configuration record ID: {config_id}")

        # Generate filename with timestamp
        filename = f"justframeit_price_export_generated_{timestamp}.xlsx"

        # Encode Excel bytes to base64 for Odoo binary field
        excel_base64 = base64.b64encode(excel_bytes).decode('ascii')

        # Prepare update values for Excel
        update_vals = {
            'x_studio_price_list_1': excel_base64,
            'x_studio_price_list_1_filename': filename,
        }

        # Only add CSV fields if CSV was generated
        if csv_bytes is not None and csv_filename is not None:
            # Encode CSV bytes to base64 for Odoo binary field
            csv_base64 = base64.b64encode(csv_bytes).decode('ascii')
            update_vals['x_studio_price_list_1_csv'] = csv_base64
            update_vals['x_studio_price_list_1_csv_filename'] = csv_filename
            logger.info("Including CSV data in update values")
        else:
            logger.info("Skipping CSV fields in update values as CSV generation was not performed")

        # Add CSV fields for each pricelist (all treated equally)
        # Note: The first CSV is already set above in update_vals, so we skip it in the loop
        additional_csv_info = []
        # First, add all CSVs to additional_csv_info for chatter message (including the first one)
        for i, (pricelist_name, csv_bytes_data, csv_filename_data) in enumerate(additional_csvs):
            csv_field_num = i + 1  # Field 1, 2, 3, etc.
            additional_csv_info.append({
                'pricelist_name': pricelist_name,
                'field': f'x_studio_price_list_{csv_field_num}_csv',
                'filename': csv_filename_data
            })

        # Now save to database, but skip the first CSV since it's already saved above
        for i, (pricelist_name, csv_bytes_data, csv_filename_data) in enumerate(additional_csvs[1:], start=1):  # Skip first CSV
            csv_field_num = i + 1  # Field 2, 3, 4, etc.
            csv_base64_data = base64.b64encode(csv_bytes_data).decode('ascii')

            update_vals[f'x_studio_price_list_{csv_field_num}_csv'] = csv_base64_data
            update_vals[f'x_studio_price_list_{csv_field_num}_csv_filename'] = csv_filename_data

            logger.info(f"Added CSV for pricelist '{pricelist_name}' to field x_studio_price_list_{csv_field_num}_csv")

        logger.info(f"Saving Excel file to x_studio_price_list_1 and {len(additional_csvs)} additional CSV files to Odoo")

        models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'write', [config_id, update_vals])

        logger.info("Excel file and CSV saved to Odoo successfully")

        # Post to chatter with summary, logs, and attachments
        logger.info("Posting to configuration chatter")

        # Create attachments for Excel, CSV files and logs
        attachment_ids = []

        # Create Excel attachment
        excel_attachment_data = {
            'name': filename,
            'type': 'binary',
            'datas': excel_base64,
            'res_model': 'x_configuration',
            'res_id': config_id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        excel_attachment_id = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'ir.attachment', 'create', [excel_attachment_data])
        attachment_ids.append(excel_attachment_id)
        logger.info(f"Created Excel attachment with ID: {excel_attachment_id}")

        # Create attachments for all CSV files (all treated equally now)
        for pricelist_name, csv_bytes_data, csv_filename_data in additional_csvs:
            csv_base64_data = base64.b64encode(csv_bytes_data).decode('ascii')
            csv_attachment_data = {
                'name': csv_filename_data,
                'type': 'binary',
                'datas': csv_base64_data,
                'res_model': 'x_configuration',
                'res_id': config_id,
                'mimetype': 'text/csv'
            }
            csv_attachment_id = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
                'ir.attachment', 'create', [csv_attachment_data])
            attachment_ids.append(csv_attachment_id)
            logger.info(f"Created CSV attachment for pricelist '{pricelist_name}' with ID: {csv_attachment_id}")

        # Get captured logs and create log attachment
        captured_logs = log_capture_string.getvalue()
        log_filename = f"price_export_logs_{timestamp}.txt"
        log_attachment_data = {
            'name': log_filename,
            'type': 'binary',
            'datas': base64.b64encode(captured_logs.encode('utf-8')).decode('ascii'),
            'res_model': 'x_configuration',
            'res_id': config_id,
            'mimetype': 'text/plain'
        }
        log_attachment_id = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'ir.attachment', 'create', [log_attachment_data])
        attachment_ids.append(log_attachment_id)
        logger.info(f"Created log attachment with ID: {log_attachment_id}")

        # Prepare response
        if csv_bytes is not None:
            csv_files_count = len(additional_csvs) + 1
            message = f'Price-export Excel and {csv_files_count} CSV files generated and saved to Odoo successfully'
        else:
            csv_files_count = 0
            message = 'Price-export Excel generated and saved to Odoo successfully (CSV generation skipped)'

        response_data = {
            'message': message,
            'config_id': config_id,
            'filename': filename,
            'csv_filename': csv_filename,
            'additional_csvs': additional_csv_info,
            'total_csv_files': csv_files_count,
            'products_processed': total_products,
            'pricelists_processed': total_pricelists,
            'source_file': source_file,
            'x_studio_is_run_locally': x_studio_is_run_locally,
            'status': 'success'
        }

        logger.info(f"Price-export generation completed - Config ID: {config_id}, Products: {total_products}, Pricelists: {total_pricelists}, CSV files: {csv_files_count}")

        # Get captured logs and clean up handler
        log_contents = log_capture_string.getvalue()
        logger.removeHandler(log_handler)
        log_capture_string.close()

        # Log the route call to Odoo logging model
        log_route_call(models, uid, '/generate-price-export', payload, log_contents, response_data)

        # Create comprehensive HTML chatter message with all logs and final return
        csv_list_items = []
        if csv_bytes is not None:
            # All CSVs are now generated equally - no "main" distinction
            for csv_info in additional_csv_info:
                csv_list_items.append(f"<li>Pricelist '{csv_info['pricelist_name']}': {csv_info['filename']}</li>")
        else:
            csv_list_items.append("<li>CSV generation was skipped (x_studio_is_run_locally = false)</li>")

        attachment_list = [f"price_export_logs_{timestamp}.txt", filename]
        if csv_bytes is not None:
            # All CSV files are in additional_csv_info now
            attachment_list.extend([csv['filename'] for csv in additional_csv_info])
        attachment_items = ", ".join(attachment_list)

        chatter_message = f"""<p><strong>✅ Price-Export Generation Completed Successfully</strong></p>


<p><strong>📝 Processing Details:</strong></p>

<p><strong>Data Retrieval:</strong></p>
<ul>
<li>Fetched {total_products} products from Odoo</li>
<li>Retrieved {total_pricelists} pricelists</li>
<li>Processed duration rules data</li>
</ul>

<p><strong>Template Processing:</strong></p>
<ul>
<li>Loaded Excel template from configuration</li>
<li>Applied product and pricing data</li>
<li>Generated main Excel file: {filename}</li>
</ul>

<p><strong>CSV Generation:</strong></p>
<ul>
<li>Created {csv_files_count} CSV files total</li>
<li>x_studio_is_run_locally: {x_studio_is_run_locally}</li>
</ul>

<p><strong>Generated CSV Files:</strong></p>
<ul>{''.join(csv_list_items)}</ul>

<p><strong>File Storage:</strong></p>
<ul>
<li>All files saved to Odoo configuration record</li>
<li>Log files created for debugging and audit trail</li>
</ul>

<p><strong>Final Return Data:</strong></p>
<ul>
<li>message: '{response_data['message']}'</li>
<li>config_id: {config_id}</li>
<li>filename: '{filename}'</li>
<li>csv_filename: '{csv_filename}'</li>
<li>total_csv_files: {csv_files_count}</li>
<li>products_processed: {total_products}</li>
<li>pricelists_processed: {total_pricelists}</li>
<li>x_studio_is_run_locally: {x_studio_is_run_locally}</li>
<li>status: 'success'</li>
</ul>"""

        # Post message to chatter
        chatter_data = {
            'body': chatter_message,
            'body_is_html': True,                # keep HTML rendering
            'message_type': 'comment',
            'subtype_xmlid': 'mail.mt_note',     # 🔑 internal note
            'attachment_ids': attachment_ids,
        }

        models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'message_post', [config_id], chatter_data)

        logger.info("Successfully posted to configuration chatter")

        return jsonify(response_data)

    except Exception as e:
        # Clean up log handler and get logs if available
        log_contents = ""
        try:
            if 'log_handler' in locals() and 'log_capture_string' in locals():
                log_contents = log_capture_string.getvalue()
                logger.removeHandler(log_handler)
                log_capture_string.close()
        except:
            pass

        logger.error(f"Error in generate-price-export route: {str(e)}")
        logger.error(f"Exception type: {type(e).__name__}")
        import traceback
        logger.error(f"Exception traceback: {traceback.format_exc()}")

        # Prepare error response data
        error_response = {'error': str(e), 'status': 'error'}

        # Log the error to Odoo logging model
        log_route_call(None, None, '/generate-price-export', payload, log_contents, error_response)

        return jsonify(error_response), 500
