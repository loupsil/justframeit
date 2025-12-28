from flask import Blueprint, jsonify, request
import xmlrpc.client
import os
import json
import base64
from io import BytesIO, StringIO
from datetime import datetime
import logging
from dotenv import load_dotenv
import pandas as pd
import openpyxl
import re
import csv
from utils import log_route_call

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Create blueprint
price_export_v2_bp = Blueprint('price-export-v2', __name__)

def create_log_capture_handler():
    """Create a string handler to capture logs during processing"""
    log_capture_string = StringIO()
    ch = logging.StreamHandler(log_capture_string)
    ch.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    ch.setFormatter(formatter)
    return ch, log_capture_string

# Odoo Configuration
    """
    Log route calls to Odoo ir.logging model.

    Args:
        models: Odoo models proxy (can be None - function will establish connection if needed)
        uid: User ID (can be None - function will establish connection if needed)
        route_name: Name of the route being called
        payload: The request payload (dict)
        server_logs: Captured server logs (string)
        response_data: The final response data returned by the route (dict)

    Returns:
        int: ID of the created log record, or None if logging failed
    """
    try:
        # If models or uid are missing, try to establish connection
        if not models or not uid:
            logger.info(f"Attempting to establish Odoo connection for logging: {route_name}")
            try:
                uid = get_uid()
                models = get_odoo_models()
                logger.info("Successfully established Odoo connection for logging")
            except Exception as conn_error:
                logger.warning(f"Cannot log to Odoo - failed to establish connection. Route: {route_name}, Error: {str(conn_error)}")
                return None

        logger.info(f"Logging route call to Odoo ir.logging model: {route_name}")

        # Format the log entry in a readable way
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Format payload as readable JSON
        payload_formatted = json.dumps(payload, indent=2, ensure_ascii=False)

        # Format response data as readable JSON
        response_formatted = json.dumps(response_data, indent=2, ensure_ascii=False)

        # Determine log level based on response status
        is_error = response_data.get('status') == 'error'
        log_level = 'error' if is_error else 'info'

        # Set appropriate name based on operation type and status
        log_name = 'Price Export Generation'

        # Create the notes content
        notes_content = f"""🕒 Timestamp: {timestamp}
📍 Route: {route_name}
⚠️  Status: {'ERROR' if is_error else 'SUCCESS'}

📋 PAYLOAD:
{payload_formatted}

📤 RESPONSE:
{response_formatted}

📝 SERVER LOGS:
{server_logs}

{'='*80}
"""

        # Create log record in Odoo ir.logging model
        log_vals = {
            'name': log_name,
            'message': response_formatted,
            'x_studio_notes': notes_content,
            'type': 'server',  # Standard type for server logs
            'level': log_level,  # Error level for failed operations, info for successful
            'path': f'Route: {route_name}',  # Required path field - set to route name
            'func': 'log_route_call',  # Function name where logging occurs
            'line': '0',  # Line number (set to 0 since we don't have actual line number)
            'dbname': ODOO_DB  # Database name from environment
        }

        # Try to create the log record in ir.logging model
        log_id = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'ir.logging', 'create', [log_vals])

        logger.info(f"Successfully created log record with ID: {log_id}")
        return log_id

    except Exception as e:
        logger.error(f"Failed to create log record in Odoo: {str(e)}")
        # Don't raise the exception - logging failure shouldn't break the main flow
        return None

# Odoo Configuration
ODOO_URL = os.getenv('JUSTFRAMEIT_ODOO_URL')
ODOO_DB = os.getenv('JUSTFRAMEIT_ODOO_DB')
ODOO_USERNAME = os.getenv('JUSTFRAMEIT_ODOO_USERNAME')
ODOO_API_KEY = os.getenv('JUSTFRAMEIT_ODOO_API_KEY')

def get_odoo_common():
    """Get Odoo common endpoint"""
    try:
        common = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/common', allow_none=True)
        return common
    except Exception as e:
        logger.error(f"Failed to connect to Odoo common endpoint: {str(e)}")
        raise

def get_odoo_models():
    """Get Odoo models endpoint"""
    try:
        models = xmlrpc.client.ServerProxy(f'{ODOO_URL}/xmlrpc/2/object', allow_none=True)
        return models
    except Exception as e:
        logger.error(f"Failed to connect to Odoo models endpoint: {str(e)}")
        raise

def get_uid():
    """Get Odoo user ID"""
    try:
        common = get_odoo_common()
        uid = common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_API_KEY, {})
        if not uid:
            raise Exception("Authentication failed")
        return uid
    except Exception as e:
        logger.error(f"Failed to authenticate with Odoo: {str(e)}")
        raise


# =============================================
# 📋 DIRECT CSV GENERATION (No Excel Calculation)
# =============================================

def get_dimensions_from_config(models, uid):
    """
    Fetch dimension configurations from Odoo.
    Returns a list of dimension tuples: [(width_mm, height_mm, width_cm, height_cm, surface_m2, circumference_m), ...]
    
    The dimensions are stored in x_configuration.x_studio_price_export_dimensions as a JSON string.
    If not found, returns default dimensions based on the Excel template.
    """
    try:
        logger.info("Fetching dimensions from Odoo configuration...")
        
        # Try to get dimensions from configuration
        config_ids = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'search', [[]]
        )
        
        if config_ids:
            config_data = models.execute_kw(
                ODOO_DB, uid, ODOO_API_KEY,
                'x_configuration', 'read', [config_ids[0]], {
                    'fields': ['x_studio_price_export_dimensions']
                }
            )
            
            if config_data and config_data[0].get('x_studio_price_export_dimensions'):
                dimensions_json = config_data[0]['x_studio_price_export_dimensions']
                logger.info(f"Found dimensions configuration in Odoo: {len(dimensions_json)} chars")
                try:
                    raw_dimensions = json.loads(dimensions_json)
                    # Convert to tuples with computed values
                    dimensions = []
                    for dim in raw_dimensions:
                        width_mm = dim['width_mm']
                        height_mm = dim['height_mm']
                        width_cm = width_mm / 10
                        height_cm = height_mm / 10
                        surface_m2 = (width_mm * height_mm) / 1000000  # mm² to m²
                        circumference_m = 2 * (width_mm + height_mm) / 1000  # mm to m
                        dimensions.append((width_mm, height_mm, width_cm, height_cm, surface_m2, circumference_m))
                    return dimensions
                except json.JSONDecodeError as e:
                    logger.warning(f"Failed to parse dimensions JSON: {e}")
        
        # Return default dimensions based on the Excel template pattern
        logger.info("Using default dimensions from Excel template pattern")
        return get_default_dimensions()
        
    except Exception as e:
        logger.warning(f"Error fetching dimensions from config: {e}")
        return get_default_dimensions()


def get_default_dimensions():
    """
    Returns the default dimension set based on the exact order from the Excel template.
    Dimensions are in the format: (width_mm, height_mm, width_cm, height_cm, surface_m2, circumference_m)
    
    The dimensions are in cm and follow the exact order as specified.
    """
    # Define dimensions in cm (width x height), in the exact order from the template
    dimension_cm_pairs = [
        # Square formats (1:1) - 18 items
        (5.0, 5.0), (70.0, 70.0), (75.0, 75.0), (80.0, 80.0), (85.0, 85.0),
        (90.0, 90.0), (95.0, 95.0), (100.0, 100.0), (105.0, 105.0), (110.0, 110.0),
        (115.0, 115.0), (120.0, 120.0), (125.0, 125.0), (130.0, 130.0), (135.0, 135.0),
        (140.0, 140.0), (145.0, 145.0), (150.0, 150.0),
        
        # Portrait 2:3 ratio - 16 items
        (5.0, 7.5), (15.0, 22.5), (25.0, 37.5), (30.0, 45.0), (35.0, 52.5),
        (45.0, 67.5), (50.0, 75.0), (55.0, 82.5), (65.0, 97.5), (70.0, 105.0),
        (76.0, 112.5), (75.0, 112.5), (85.0, 127.5), (90.0, 135.0), (95.0, 142.5),
        (100.0, 150.0),
        
        # Portrait ~3:4 ratio - 22 items
        (5.0, 7.0), (10.0, 13.0), (20.0, 27.0), (25.0, 33.0), (35.0, 47.0),
        (40.0, 53.0), (40.0, 60.0), (40.0, 53.0), (45.0, 60.0), (50.0, 67.0),
        (55.0, 73.0), (65.0, 87.0), (70.0, 93.0), (75.0, 100.0), (80.0, 107.0),
        (85.0, 113.0), (90.0, 120.0), (95.0, 127.0), (100.0, 133.0), (105.0, 140.0),
        (110.0, 147.0), (115.0, 153.0),
        
        # Landscape 3:2 ratio - 20 items
        (7.5, 5.0), (15.0, 10.0), (22.5, 15.0), (30.0, 20.0), (37.5, 25.0),
        (45.0, 30.0), (52.5, 35.0), (60.0, 40.0), (67.5, 45.0), (75.0, 50.0),
        (82.5, 55.0), (90.0, 60.0), (97.5, 65.0), (105.0, 70.0), (112.5, 75.0),
        (120.0, 80.0), (127.5, 85.0), (135.0, 90.0), (142.5, 95.0), (150.0, 100.0),
        
        # Landscape ~4:3 ratio - 23 items
        (6.7, 5.0), (13.3, 10.0), (20.0, 15.0), (26.7, 20.0), (33.3, 25.0),
        (40.0, 30.0), (46.7, 35.0), (53.3, 40.0), (60.0, 45.0), (66.0, 50.0),
        (73.0, 55.0), (80.0, 60.0), (86.0, 65.0), (93.0, 70.0), (100.0, 75.0),
        (106.0, 80.0), (113.0, 85.0), (120.0, 90.0), (126.0, 95.0), (133.0, 100.0),
        (140.0, 105.0), (146.0, 110.0), (153.0, 115.0),
        
        # Standard photo formats - 12 items
        (9.0, 13.0), (20.0, 28.0), (18.0, 13.0), (24.0, 18.0), (25.0, 20.0),
        (28.0, 20.0), (29.7, 21.0), (30.0, 24.0), (50.0, 40.0), (60.0, 50.0),
        (70.0, 50.0), (100.0, 70.0),
        
        # Mixed formats - 34 items
        (13.0, 9.0), (10.0, 10.0), (10.5, 10.5), (10.0, 15.0), (13.0, 13.0),
        (25.0, 25.0), (13.0, 18.0), (15.0, 20.0), (15.0, 15.0), (35.0, 35.0),
        (18.0, 24.0), (20.0, 20.0), (45.0, 45.0), (20.0, 25.0), (55.0, 55.0),
        (20.0, 30.0), (21.0, 29.7), (60.0, 60.0), (65.0, 65.0), (70.0, 70.0),
        (24.0, 30.0), (30.0, 30.0), (30.0, 40.0), (30.0, 45.0), (40.0, 40.0),
        (40.0, 50.0), (40.0, 60.0), (50.0, 50.0), (50.0, 60.0), (50.0, 70.0),
        (60.0, 80.0), (60.0, 90.0), (70.0, 100.0), (80.0, 120.0),
    ]
    
    dimensions = []
    for width_cm, height_cm in dimension_cm_pairs:
        width_mm = width_cm * 10
        height_mm = height_cm * 10
        surface_m2 = (width_mm * height_mm) / 1000000  # mm² to m²
        circumference_m = 2 * (width_mm + height_mm) / 1000  # mm to m
        dimensions.append((width_mm, height_mm, width_cm, height_cm, surface_m2, circumference_m))
    
    logger.info(f"Generated {len(dimensions)} default dimensions")
    return dimensions


def build_duration_lookup(duration_rules):
    """
    Pre-build a lookup dictionary for duration rules, organized by service name.
    Each service has a sorted list of (quantity, duration) tuples for fast binary search.
    
    Args:
        duration_rules: List of duration rule dictionaries from Odoo
        
    Returns:
        dict: {service_name: {'sorted_rules': [(qty, duration), ...], 'max_duration': duration}}
    """
    import bisect
    
    lookup = {}
    
    for rule in duration_rules:
        service_name = get_service_name(rule.get('x_associated_service'))
        if not service_name:
            continue
            
        qty = rule.get('x_studio_quantity') or 0
        duration = rule.get('x_duurtijd_totaal') or 0
        
        if service_name not in lookup:
            lookup[service_name] = {'rules': [], 'max_duration': 0, 'max_qty': 0}
        
        lookup[service_name]['rules'].append((qty, duration))
        
        # Track the rule with max quantity for fallback
        if qty > lookup[service_name]['max_qty']:
            lookup[service_name]['max_qty'] = qty
            lookup[service_name]['max_duration'] = duration
    
    # Sort rules by quantity for each service (enables binary search)
    for service_name in lookup:
        lookup[service_name]['rules'].sort(key=lambda x: x[0])
        # Extract just quantities for bisect
        lookup[service_name]['quantities'] = [r[0] for r in lookup[service_name]['rules']]
        lookup[service_name]['durations'] = [r[1] for r in lookup[service_name]['rules']]
    
    return lookup


def lookup_service_duration_fast(service_name, quantity_threshold, duration_lookup):
    """
    Fast duration lookup using pre-built dictionary and binary search.
    
    Args:
        service_name: The service name to look up
        quantity_threshold: The dimension value to match
        duration_lookup: Pre-built lookup dictionary from build_duration_lookup()
        
    Returns:
        The duration in seconds, or 0 if not found
    """
    import bisect
    
    if service_name not in duration_lookup:
        return 0
    
    service_data = duration_lookup[service_name]
    quantities = service_data['quantities']
    durations = service_data['durations']
    
    if not quantities:
        return 0
    
    # Binary search: find leftmost quantity >= threshold
    idx = bisect.bisect_left(quantities, quantity_threshold)
    
    if idx < len(quantities):
        # Found a quantity >= threshold
        return durations[idx]
    else:
        # No matching rule, return max quantity rule's duration (fallback)
        return service_data['max_duration']


def lookup_service_duration(service_name, quantity_threshold, duration_rules):
    """
    Replicate the Excel MINIFS/FILTER/INDEX logic to find service duration.
    
    Excel formula logic:
    MINIFS('Service Duration'!$C$2:$C$401, 'Service Duration'!$A$2:$A$401, $H7, 
           'Service Duration'!$C$2:$C$401, ">=" & dimension_value)
    
    Then INDEX/FILTER to get the duration for that minimum quantity.
    
    Args:
        service_name: The service name to look up (e.g., "Framing")
        quantity_threshold: The dimension value (surface or circumference) to match
        duration_rules: List of duration rule dictionaries from Odoo
        
    Returns:
        The duration in seconds, or 0 if not found
    """
    # Filter rules for this service
    service_rules = [
        r for r in duration_rules 
        if get_service_name(r.get('x_associated_service')) == service_name
    ]
    
    if not service_rules:
        return 0
    
    # Find the minimum quantity that is >= the threshold
    # Filter rules where quantity >= threshold
    matching_rules = [
        r for r in service_rules 
        if (r.get('x_studio_quantity') or 0) >= quantity_threshold
    ]
    
    if not matching_rules:
        # If no matching rules, find the highest quantity rule as fallback
        max_qty_rule = max(service_rules, key=lambda r: r.get('x_studio_quantity') or 0, default=None)
        if max_qty_rule:
            return max_qty_rule.get('x_duurtijd_totaal') or 0
        return 0
    
    # Find the rule with minimum quantity among matching rules
    min_qty_rule = min(matching_rules, key=lambda r: r.get('x_studio_quantity') or 0)
    
    return min_qty_rule.get('x_duurtijd_totaal') or 0


def get_service_name(service_value):
    """Extract service name from Odoo field value (handles tuple/list format)"""
    if isinstance(service_value, (list, tuple)):
        return str(service_value[1]) if len(service_value) > 1 else str(service_value[0]) if service_value else ''
    return str(service_value) if service_value else ''


def compute_prices_vectorized(product, dimensions, duration_lookup, margin):
    """
    Compute prices for a product across ALL dimensions at once using vectorized operations.
    
    This is much faster than calling compute_price individually for each dimension.
    
    Args:
        product: Product dictionary from Odoo
        dimensions: List of tuples (width_mm, height_mm, width_cm, height_cm, surface_m2, circumference_m)
        duration_lookup: Pre-built lookup dictionary from build_duration_lookup()
        margin: Margin/markup percentage (e.g., 0.5 for 50%)
        
    Returns:
        List of computed prices (floats)
    """
    import numpy as np
    
    # Pre-extract product fields (once per product, not per dimension)
    price_computation = get_service_name(product.get('x_studio_price_computation'))
    standard_price = product.get('standard_price') or 0
    service_name = get_service_name(product.get('x_studio_associated_service'))
    cost_per_hour = product.get('x_studio_associated_cost_per_employee_per_hour') or 0
    
    is_circumference = (price_computation == 'Circumference')
    
    # Extract dimension arrays (vectorized)
    surfaces = np.array([d[4] for d in dimensions])  # surface_m2
    circumferences = np.array([d[5] for d in dimensions])  # circumference_m
    
    # Determine dimension values and base costs based on computation method
    if is_circumference:
        dimension_values = circumferences
        base_costs = circumferences * standard_price
    else:
        dimension_values = surfaces
        base_costs = surfaces * standard_price
    
    # Lookup durations for all dimensions (this is still per-dimension but uses fast lookup)
    durations = np.array([
        lookup_service_duration_fast(service_name, dv, duration_lookup)
        for dv in dimension_values
    ])
    
    # Calculate labor costs: duration * cost_per_hour / 3600
    labor_costs = (durations * cost_per_hour) / 3600
    
    # Apply margin: (base + labor) * (1 + margin)
    total_prices = (base_costs + labor_costs) * (1 + margin)
    
    # Round to 2 decimal places
    return np.round(total_prices, 2).tolist()


def compute_price(product, dimension, duration_rules, margin):
    """
    Compute the price for a product at a specific dimension.
    
    Replicates the Excel formula:
    =(IF($F7=$K$4, L$4*$G7, L$3*$G7) + 
      (INDEX(FILTER('Service Duration'!$D$2:$D$401, ...)) * $J7/3600)
     ) * (1 + $K7)
    
    Where:
    - $F7 = price computation method (Surface or Circumference)
    - $K$4 = "Circumference" (constant)
    - L$4 = circumference value for this dimension (in meters)
    - L$3 = surface value for this dimension (in m²)
    - $G7 = standard_price (cost per unit)
    - $H7 = associated service name
    - $J7 = cost per employee per hour
    - $K7 = margin (from pricelist)
    
    Args:
        product: Product dictionary from Odoo
        dimension: Tuple (width_mm, height_mm, width_cm, height_cm, surface_m2, circumference_m)
        duration_rules: List of duration rule dictionaries from Odoo
        margin: Margin/markup percentage (e.g., 0.5 for 50%)
        
    Returns:
        Computed price (float)
    """
    width_mm, height_mm, width_cm, height_cm, surface_m2, circumference_m = dimension
    
    # Get product fields
    price_computation = get_service_name(product.get('x_studio_price_computation'))
    standard_price = product.get('standard_price') or 0
    service_name = get_service_name(product.get('x_studio_associated_service'))
    cost_per_hour = product.get('x_studio_associated_cost_per_employee_per_hour') or 0
    
    # Determine base cost based on price computation method
    # IF($F7=$K$4, L$4*$G7, L$3*$G7)
    if price_computation == 'Circumference':
        base_cost = circumference_m * standard_price
        dimension_value = circumference_m
    else:  # Surface
        base_cost = surface_m2 * standard_price
        dimension_value = surface_m2
    
    # Lookup service duration
    duration_seconds = lookup_service_duration(service_name, dimension_value, duration_rules)
    
    # Calculate labor cost: duration * cost_per_hour / 3600
    labor_cost = (duration_seconds * cost_per_hour) / 3600
    
    # Apply margin: (base + labor) * (1 + margin)
    total_price = (base_cost + labor_cost) * (1 + margin)
    
    return round(total_price, 2)


def generate_csv_direct(models, uid, pricelist_name=None):
    """
    Generate CSV directly by computing prices in Python without Excel calculation.
    
    Args:
        models: Odoo models proxy
        uid: Odoo user ID
        pricelist_name: Optional specific pricelist name to generate CSV for.
                       If None, generates CSVs for all pricelists.
    
    Returns:
        If pricelist_name is specified: (csv_bytes, pricelist_name, csv_filename)
        If pricelist_name is None: list of (csv_bytes, pricelist_name, csv_filename) tuples
    """
    try:
        logger.info(f"Starting direct CSV generation (pricelist: {pricelist_name or 'ALL'})")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # =============================================
        # 📋 STEP 1: Fetch products with price computation = Surface or Circumference
        # =============================================
        logger.info("Fetching products with price computation = Surface or Circumference...")
        
        products = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'product.product', 'search_read',
            [[['x_studio_price_computation', 'in', ['Surface', 'Circumference']]]],
            {
                'fields': [
                    'name',
                    'id',
                    'product_tmpl_id',
                    'x_studio_product_code',
                    'x_studio_location_code',
                    'description_ecommerce',
                    'x_studio_price_computation',
                    'standard_price',
                    'x_studio_associated_service',
                    'x_studio_associated_work_center',
                    'x_studio_associated_cost_per_employee_per_hour'
                ]
            }
        )
        
        logger.info(f"Fetched {len(products)} products")
        
        # =============================================
        # 📋 STEP 2: Fetch pricelists
        # =============================================
        logger.info("Fetching pricelists...")
        
        all_pricelists = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'product.pricelist', 'search_read',
            [],
            {
                'fields': ['name', 'x_studio_price_discount'],
                'limit': 100
            }
        )
        
        # Filter out "Default" pricelist
        pricelists = [
            p for p in all_pricelists
            if p.get('name', '').lower() != 'default'
        ]
        
        # If specific pricelist requested, filter to just that one
        if pricelist_name:
            pricelists = [p for p in pricelists if p.get('name') == pricelist_name]
            if not pricelists:
                raise Exception(f"Pricelist '{pricelist_name}' not found")
        
        logger.info(f"Processing {len(pricelists)} pricelists")
        
        # =============================================
        # 📋 STEP 3: Fetch service duration rules
        # =============================================
        logger.info("Fetching service duration rules...")
        
        duration_rules = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'x_services_duration_rules', 'search_read',
            [[]],
            {
                'fields': [
                    'x_associated_service',
                    'x_studio_work_center',
                    'x_studio_quantity',
                    'x_duurtijd_totaal'
                ]
            }
        )
        
        logger.info(f"Fetched {len(duration_rules)} duration rules")
        
        # =============================================
        # 📋 STEP 4: Get dimensions and build lookup
        # =============================================
        dimensions = get_dimensions_from_config(models, uid)
        logger.info(f"Using {len(dimensions)} dimensions")
        
        # Pre-build duration lookup for fast access (O(1) instead of O(n) per lookup)
        logger.info("Building duration rules lookup dictionary...")
        duration_lookup = build_duration_lookup(duration_rules)
        logger.info(f"Built lookup for {len(duration_lookup)} services")
        
        # Pre-compute dimension labels once
        dimension_labels = [f"{dim[2]} x {dim[3]}" for dim in dimensions]
        
        # =============================================
        # 📋 STEP 5: Generate CSV for each pricelist
        # =============================================
        csv_results = []
        
        for pricelist in pricelists:
            pl_name = pricelist.get('name', 'Unknown')
            raw_discount = pricelist.get('x_studio_price_discount') or 0
            
            # Apply the Excel formula: (x_studio_price_discount * -1) / 100
            # Example: if x_studio_price_discount = -50, margin = (-50 * -1) / 100 = 0.5 (50%)
            margin = (raw_discount * -1) / 100
            
            logger.info(f"Generating CSV for pricelist '{pl_name}' with raw_discount={raw_discount}, margin={margin}")
            
            # Create CSV data with comma delimiter for proper Excel column separation
            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer, delimiter=',')
            
            # Create header row with dimension labels
            header_row = ['product_tmpl_id'] + dimension_labels
            writer.writerow(header_row)
            
            # Create data rows for each product using vectorized computation
            for idx, product in enumerate(products):
                # Get product template ID
                product_tmpl_id = product.get('product_tmpl_id')
                if isinstance(product_tmpl_id, (list, tuple)):
                    product_tmpl_id = product_tmpl_id[0]
                
                # Compute prices for ALL dimensions at once (vectorized)
                prices = compute_prices_vectorized(product, dimensions, duration_lookup, margin)
                
                # Format as strings
                price_strings = [f"{p:.2f}" for p in prices]
                
                row = [str(product_tmpl_id)] + price_strings
                writer.writerow(row)
                
                # Log progress every 500 products
                if (idx + 1) % 500 == 0:
                    logger.info(f"Processed {idx + 1}/{len(products)} products for '{pl_name}'")
            
            # Get CSV bytes
            csv_content = csv_buffer.getvalue()
            csv_bytes = csv_content.encode('utf-8')
            
            # Generate filename
            safe_name = pl_name.replace(' ', '_').lower()
            csv_filename = f"justframeit_price_export_{safe_name}_{timestamp}.csv"
            
            csv_results.append((csv_bytes, pl_name, csv_filename))
            logger.info(f"Generated CSV for '{pl_name}': {len(products)} products x {len(dimensions)} dimensions")
        
        # Return results
        if pricelist_name:
            return csv_results[0] if csv_results else None
        else:
            return csv_results
            
    except Exception as e:
        logger.error(f"Error in direct CSV generation: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise


def generate_price_export_excel(models, uid, d3_formula=None):
    """
    Generate Excel file using the exact same logic as the Jupyter notebook.
    Creates a new Excel file with 3 tabs populated with Odoo data.
    Returns the Excel file as bytes, total products, total pricelists, and total duration rules.

    Args:
        models: Odoo models proxy
        uid: Odoo user ID
        d3_formula: Optional formula to set in tab 2, cell D3 (e.g., '=B4', '=B5', etc.)
    """
    try:
        logger.info("Starting Excel price-export generation using exact Jupyter notebook logic")

        import time
        import shutil
        from openpyxl import load_workbook

        # =============================================
        # 📋 STEP 1: Fetch products with price computation = Surface or Circumference
        # =============================================
        logger.info("Fetching products with price computation = Surface or Circumference...")

        products = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'product.product', 'search_read',
            [[['x_studio_price_computation', 'in', ['Surface', 'Circumference']]]],  # Filter by price computation
            {
                'fields': [
                    'name',
                    'id',
                    'x_studio_product_code',
                    'x_studio_location_code',
                    'description_ecommerce',
                    'x_studio_price_computation',
                    'standard_price',
                    'x_studio_associated_service',
                    'x_studio_associated_work_center',
                    'x_studio_associated_cost_per_employee_per_hour'
                ],
                # uncomment for testing with a limit
                #'limit': 100
            }
        )

        total_products = len(products)
        logger.info(f"Fetched {total_products} products")

        # =============================================
        # 📋 STEP 2: Create new Excel file from template
        # =============================================
        # Fetch template from Odoo x_configuration.x_studio_price_export_template field
        logger.info("Fetching Excel template from Odoo x_configuration.x_studio_price_export_template")
        config_ids = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'search', [[]]
        )

        if not config_ids:
            raise Exception("No x_configuration record found")

        config_id = config_ids[0]
        config_data = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'read', [config_id], {
                'fields': ['x_studio_price_export_template']
            }
        )

        if not config_data or not config_data[0].get('x_studio_price_export_template'):
            raise Exception("No template found in x_configuration.x_studio_price_export_template field")

        # Decode the base64 template data
        template_base64 = config_data[0]['x_studio_price_export_template']
        logger.info(f"Raw template base64 data length: {len(template_base64) if template_base64 else 0} characters")

        if not template_base64:
            raise Exception("Template data from Odoo is empty")

        try:
            template_bytes = base64.b64decode(template_base64)
            logger.info(f"Decoded template bytes length: {len(template_bytes)} bytes")
        except Exception as e:
            raise Exception(f"Failed to decode template base64 data from Odoo: {str(e)}")

        if len(template_bytes) < 100:  # Basic sanity check for Excel file size
            raise Exception(f"Template data from Odoo is suspiciously small ({len(template_bytes)} bytes), likely corrupted")

        # Generate timestamp for filenames
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Save template to temporary file
        template_file = f'temp_template_{timestamp}.xlsx'
        with open(template_file, 'wb') as f:
            f.write(template_bytes)
        logger.info(f"Saved template from Odoo to temporary file: {template_file} ({len(template_bytes)} bytes)")

        # Validate the template file can be loaded
        try:
            test_wb = load_workbook(template_file, read_only=True)
            test_wb.close()
            logger.info("Template file validation successful")
        except Exception as e:
            # Clean up the corrupted file
            try:
                os.remove(template_file)
            except:
                pass
            raise Exception(f"Template file from Odoo is corrupted and cannot be loaded: {str(e)}")
        output_file = f'justframeit_pricelist_{timestamp}.xlsx'

        # Copy the template to create a new file
        shutil.copy(template_file, output_file)
        logger.info(f"Created new file: {output_file}")

        # Load the newly created Excel file
        wb = load_workbook(output_file)

        # =============================================
        # 📋 STEP 2.5: Modify D3 formula in TAB 2 if specified
        # =============================================
        if d3_formula is not None:
            logger.info(f"Modifying D3 formula in tab 2 to: {d3_formula}")
            if len(wb.sheetnames) > 1:
                ws2 = wb.worksheets[1]
                ws2['D3'] = d3_formula
                logger.info("D3 formula modified successfully")
            else:
                logger.warning("Could not modify D3 formula - second tab not found")

        # =============================================
        # 📋 STEP 3: Fill TAB 1 with products
        # =============================================
        logger.info("Filling TAB 1 with products data...")
        ws1 = wb.worksheets[0]  # First tab (index 0)

        # Starting row is A7 (row 7)
        start_row = 7
        start_time = time.time()

        # Fill data for each product
        for idx, product in enumerate(products):
            current_row = start_row + idx

            # Helper function to convert list/tuple values to string
            def convert_value(value):
                if isinstance(value, (list, tuple)):
                    # If it's a list/tuple, join elements or take the second element (name)
                    if len(value) > 1:
                        return str(value[1])  # Return the name part
                    return str(value[0]) if value else ''
                return value

            # Fill columns A through J with the product data
            ws1[f'A{current_row}'] = convert_value(product.get('name'))
            ws1[f'B{current_row}'] = convert_value(product.get('id'))
            ws1[f'C{current_row}'] = convert_value(product.get('x_studio_product_code'))
            ws1[f'D{current_row}'] = convert_value(product.get('x_studio_location_code'))
            ws1[f'E{current_row}'] = convert_value(product.get('description_ecommerce'))
            ws1[f'F{current_row}'] = convert_value(product.get('x_studio_price_computation'))
            ws1[f'G{current_row}'] = convert_value(product.get('standard_price'))
            ws1[f'H{current_row}'] = convert_value(product.get('x_studio_associated_service'))
            ws1[f'I{current_row}'] = convert_value(product.get('x_studio_associated_work_center'))
            ws1[f'J{current_row}'] = convert_value(product.get('x_studio_associated_cost_per_employee_per_hour'))

            if idx % 50 == 0:  # Log progress every 50 products
                logger.info(f"Filled row {current_row} on tab 1 with product: {product.get('name')}")

        # Save the workbook
        wb.save(output_file)
        logger.info(f"Successfully filled {len(products)} products into tab 1")

        # =============================================
        # 📋 STEP 4: Fetch pricelists and fill TAB 2
        # =============================================
        logger.info("Fetching pricelists and filling TAB 2...")

        # Fetch all pricelists and filter out "Default" in Python to avoid domain parsing issues
        all_pricelists = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'product.pricelist', 'search_read',
            [],  # Empty domain = get all pricelists
            {
                'fields': ['name', 'x_studio_price_discount'],
                'limit': 100  # Add limit to prevent excessive results
            }
        )

        # Filter out pricelists named "Default" (case insensitive)
        pricelists = [
            p for p in all_pricelists
            if p.get('name', '').lower() != 'default'
        ]

        total_pricelists = len(pricelists)
        logger.info(f"Fetched {total_pricelists} pricelists")

        # Access the second tab (sheet)
        if len(wb.sheetnames) > 1:
            ws2 = wb.worksheets[1]
        else:
            logger.warning("Second tab not found in workbook")
            ws2 = None

        if ws2:
            # Starting row is A3 (row 3)
            pricelist_start_row = 3

            # Fill data for each pricelist
            for idx, pricelist in enumerate(pricelists):
                current_row = pricelist_start_row + idx

                # Helper function to convert list/tuple values to string
                def convert_value(value):
                    if isinstance(value, (list, tuple)):
                        if len(value) > 1:
                            return str(value[1])
                        return str(value[0]) if value else ''
                    return value

                # Fill columns A and B with pricelist data
                ws2[f'A{current_row}'] = convert_value(pricelist.get('name'))
                ws2[f'B{current_row}'] = convert_value(pricelist.get('x_studio_price_discount'))

                if idx % 50 == 0:  # Log progress every 50 pricelists
                    logger.info(f"Filled row {current_row} on tab 2 with pricelist: {pricelist.get('name')}")

            # Save the workbook again with pricelist data
            wb.save(output_file)
            logger.info(f"Successfully filled {len(pricelists)} pricelists into tab 2")

        # =============================================
        # 📋 STEP 5: Fetch service duration rules and fill TAB 3
        # =============================================
        logger.info("Fetching service duration rules and filling TAB 3...")

        duration_rules = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'x_services_duration_rules', 'search_read',
            [[]],  # Empty domain = all duration rules
            {
                'fields': [
                    'x_associated_service',
                    'x_studio_work_center',
                    'x_studio_quantity',
                    'x_duurtijd_totaal'
                ]
            }
        )

        total_duration_rules = len(duration_rules)
        logger.info(f"Fetched {total_duration_rules} service duration rules")

        # Access the third tab (sheet)
        if len(wb.sheetnames) > 2:
            ws3 = wb.worksheets[2]
        else:
            logger.warning("Third tab not found in workbook")
            ws3 = None

        if ws3:
            # Starting row is A3 (row 3)
            duration_start_row = 3

            # Fill data for each duration rule
            for idx, rule in enumerate(duration_rules):
                current_row = duration_start_row + idx

                # Helper function to convert list/tuple values to string
                def convert_value(value):
                    if isinstance(value, (list, tuple)):
                        if len(value) > 1:
                            return str(value[1])
                        return str(value[0]) if value else ''
                    return value

                # Fill columns A through D with duration rule data
                ws3[f'A{current_row}'] = convert_value(rule.get('x_associated_service'))
                ws3[f'B{current_row}'] = convert_value(rule.get('x_studio_work_center'))
                ws3[f'C{current_row}'] = convert_value(rule.get('x_studio_quantity'))
                ws3[f'D{current_row}'] = convert_value(rule.get('x_duurtijd_totaal'))

                if idx % 50 == 0:  # Log progress every 50 rules
                    logger.info(f"Filled row {current_row} on tab 3 with service: {rule.get('x_associated_service')}")

            # Save the workbook again with duration rules data
            wb.save(output_file)
            logger.info(f"Successfully filled {len(duration_rules)} service duration rules into tab 3")

        # =============================================
        # 📋 STEP 6: Read the final Excel file and return bytes
        # =============================================
        logger.info("Reading final Excel file to return as bytes...")

        # Read the file back as bytes
        with open(output_file, 'rb') as f:
            excel_bytes = f.read()

        # Clean up temporary files
        try:
            os.remove(template_file)
            logger.info(f"Cleaned up temporary template file: {template_file}")
        except Exception as e:
            logger.warning(f"Failed to clean up temporary template file {template_file}: {str(e)}")

        try:
            os.remove(output_file)
            logger.info(f"Cleaned up temporary output file: {output_file}")
        except Exception as e:
            logger.warning(f"Failed to clean up temporary output file {output_file}: {str(e)}")

        logger.info("Excel file generated successfully with all tabs populated")

        # Return: excel_bytes, products_count, pricelists_count, duration_rules_count, source_description
        return excel_bytes, total_products, total_pricelists, f"Generated from Odoo data: {total_products} products, {total_pricelists} pricelists, {total_duration_rules} duration rules"

    except Exception as e:
        logger.error(f"Error generating price-export Excel: {str(e)}")
        raise

def generate_csv_from_excel(excel_bytes):
    """
    Generate CSV from Excel bytes using the exact same logic as the Jupyter notebook.
    Uses xlwings to force calculations before extracting data.
    Takes Excel bytes, processes the first worksheet, and returns CSV bytes.
    """
    try:
        logger.info("Starting CSV generation from Excel bytes using xlwings calculation")

        import xlwings as xw
        import tempfile

        # Save Excel bytes to temporary file for xlwings
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_excel:
            temp_excel.write(excel_bytes)
            temp_excel_path = temp_excel.name

        logger.info(f"Saved Excel bytes to temporary file: {temp_excel_path}")

        try:
            # Open with xlwings to force calculation
            logger.info("Opening Excel file with xlwings to force calculations")
            app = xw.App(visible=False)
            wb_xlwings = app.books.open(temp_excel_path)
            ws_xlwings = wb_xlwings.sheets[0]

            # Force calculation
            logger.info("Forcing Excel calculation...")
            wb_xlwings.app.calculate()

            # Save the calculated workbook to another temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_calculated:
                temp_calculated_path = temp_calculated.name

            wb_xlwings.save(temp_calculated_path)
            wb_xlwings.close()
            app.quit()

            logger.info(f"Calculated workbook saved as: {temp_calculated_path}")

            # Now load with openpyxl
            logger.info("Loading calculated Excel file with openpyxl")
            wb = openpyxl.load_workbook(temp_calculated_path, data_only=True)
            logger.info("Calculated Excel file loaded successfully")

        finally:
            # Clean up the first temporary file
            try:
                os.remove(temp_excel_path)
                logger.info(f"Cleaned up temporary input file: {temp_excel_path}")
            except Exception as e:
                logger.warning(f"Could not remove temporary input file: {e}")

        # Access the first tab
        ws = wb.worksheets[0]
        logger.info(f"Accessing first worksheet: {ws.title}")

        # Headers are in row 6
        header_row = 6
        start_row = 7  # Data starts from row 7

        # Find the last row with data in column A (starting from row 7)
        last_row = start_row

        logger.info("Finding last row with data...")
        # Find where column A becomes 0 or empty
        for row in range(start_row, ws.max_row + 1):
            if row % 100 == 0:  # Log progress every 100 rows
                logger.info(f"   Checking row {row}...")
            cell_value = ws[f'A{row}'].value
            if cell_value == 0 or cell_value is None or cell_value == '':
                break
            last_row = row

        logger.info(f"Extracting data from row {start_row} to row {last_row}")

        # Get headers from row 6, but check row 5 for dimension indicators
        logger.info(f"Reading headers from row {header_row}...")
        headers = []
        valid_columns = []  # Track which columns to include
        dimension_columns = []  # Track which columns are dimensions
        found_first_dimension = False
        last_dimension_col = 0
        for col in range(1, ws.max_column + 1):
            if col % 50 == 0:  # Log progress every 50 columns
                logger.info(f"   Processing column {col}...")
            # Check row 5 for dimension indicator (e.g., "10.0 x 20.0")
            dimension_cell = ws.cell(row=5, column=col)
            dimension_value = dimension_cell.value

            # Get the header from row 6
            header_cell = ws.cell(row=header_row, column=col)
            header_value = header_cell.value

            # Stricter check: dimension must contain 'x' AND at least one number
            is_dimension = False
            if dimension_value and isinstance(dimension_value, str):
                # Check if it contains 'x' and at least one digit
                if 'x' in dimension_value.lower() and re.search(r'\d', dimension_value):
                    is_dimension = True
                    found_first_dimension = True
                    last_dimension_col = col

            # If we found dimensions before but now stopped finding them, stop processing columns
            if found_first_dimension and not is_dimension:
                logger.info(f"Stopping at column {col} - no more dimension headers found")
                break

            # If row 5 contains a valid dimension indicator, use it
            if is_dimension:
                headers.append(dimension_value)
                valid_columns.append(col)
                dimension_columns.append(len(headers) - 1)  # Track the index of this dimension column
            elif header_value and not found_first_dimension:
                # Only include non-dimension columns if we haven't found dimensions yet
                headers.append(header_value)
                valid_columns.append(col)
            elif not found_first_dimension:
                headers.append(f'Column_{col}')
                valid_columns.append(col)

        logger.info(f"Found {len(headers)} columns (last dimension column: {last_dimension_col})")

        # Extract data from row 7 onwards to the last valid row
        logger.info(f"Extracting data from {last_row - start_row + 1} rows...")
        data = []
        for row in range(start_row, last_row + 1):
            if (row - start_row) % 50 == 0:  # Log progress every 50 rows
                logger.info(f"   Processing row {row} ({row - start_row + 1}/{last_row - start_row + 1})...")
            row_data = []
            for idx, col in enumerate(valid_columns):  # Only extract data from valid columns
                cell = ws.cell(row=row, column=col)
                # Get the actual value, not the formula
                cell_value = cell.value

                # Check if this is a dimension column
                is_dimension_col = idx in dimension_columns

                # Handle special cases
                if cell_value is None:
                    row_data.append(None)
                elif isinstance(cell_value, str) and cell_value.startswith('!REF'):
                    # Handle reference errors
                    row_data.append(None)
                elif isinstance(cell_value, (int, float)):
                    # Round all numeric values to 2 decimal places
                    row_data.append(round(float(cell_value), 2))
                else:
                    row_data.append(cell_value)
            data.append(row_data)

        # Create DataFrame
        logger.info("Creating DataFrame...")
        df = pd.DataFrame(data, columns=headers)
        logger.info(f"DataFrame created with {len(df)} rows and {len(df.columns)} columns")

        # Convert DataFrame to CSV bytes
        logger.info("Converting DataFrame to CSV bytes...")
        csv_buffer = BytesIO()
        df.to_csv(csv_buffer, index=False)
        csv_bytes = csv_buffer.getvalue()

        logger.info(f"Successfully generated CSV with {len(df)} rows from Excel data")

        # Clean up the calculated temporary file
        try:
            os.remove(temp_calculated_path)
            logger.info(f"Cleaned up temporary calculated file: {temp_calculated_path}")
        except Exception as e:
            logger.warning(f"Could not remove temporary calculated file: {e}")

        return csv_bytes

    except Exception as e:
        logger.error(f"Error generating CSV from Excel: {str(e)}")
        raise


def process_pricelist_for_parallel(pricelist_name, pricelist_row, timestamp):
    """
    Process a single pricelist to generate its Excel file.
    This function must be at module level to be pickleable for multiprocessing.
    Establishes its own Odoo connection since connections cannot be pickled.
    Returns Excel bytes - CSV conversion happens in main process due to xlwings limitations.
    """
    try:
        logger.info(f"Processing pricelist '{pricelist_name}' at row {pricelist_row}")

        # Calculate the formula for D3 (pointing to column B of this pricelist row)
        d3_formula = f"=B{pricelist_row}"

        # Establish new Odoo connection for this worker process
        logger.info(f"Establishing Odoo connection for pricelist '{pricelist_name}'")
        worker_uid = get_uid()
        worker_models = get_odoo_models()
        logger.info(f"Successfully connected to Odoo for pricelist '{pricelist_name}'")

        # Add retry logic for Excel generation to handle transient errors
        max_retries = 3
        for attempt in range(max_retries):
            try:
                logger.info(f"Generating Excel for pricelist '{pricelist_name}' (attempt {attempt + 1}/{max_retries})")
                # Generate Excel with modified D3 formula
                modified_excel_bytes, _, _, _ = generate_price_export_excel(worker_models, worker_uid, d3_formula=d3_formula)
                logger.info(f"Successfully generated Excel for pricelist '{pricelist_name}' ({len(modified_excel_bytes)} bytes)")
                break  # Success, exit retry loop
            except Exception as e:
                logger.warning(f"Attempt {attempt + 1} failed for pricelist '{pricelist_name}': {str(e)}")
                if attempt == max_retries - 1:  # Last attempt
                    raise
                # Wait before retry
                import time
                time.sleep(1)

        # Return Excel bytes - CSV conversion will happen in main process
        logger.info(f"Successfully generated Excel for pricelist '{pricelist_name}' - CSV conversion will happen in main process")

        return (pricelist_name, modified_excel_bytes, timestamp)

    except Exception as e:
        logger.error(f"Error processing pricelist '{pricelist_name}': {str(e)}")
        logger.error(f"Exception type: {type(e).__name__}")
        import traceback
        logger.error(f"Exception traceback: {traceback.format_exc()}")
        raise

def generate_csvs_from_pricelists(models, uid, original_excel_bytes, timestamp):
    """
    Generate CSV files by modifying D3 formula to point to different pricelists.
    All pricelists are processed using the same logic with parallelization.

    This function:
    1. Loads the original Excel and reads pricelist names from tab 2, column A starting from row 3
    2. For each pricelist, modifies D3 to point to that row's B column
    3. Generates CSV from each modified Excel using parallel processing
    4. Returns list of (pricelist_name, csv_bytes, csv_filename) tuples

    Args:
        models: Odoo models proxy
        uid: Odoo user ID
        original_excel_bytes: The original Excel file bytes
        timestamp: Timestamp string for filename generation

    Returns:
        list: List of tuples (pricelist_name, csv_bytes, csv_filename)
    """
    try:
        logger.info("Starting generation of CSVs from pricelists")

        import tempfile

        # Save original Excel bytes to temporary file for reading pricelist names
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_excel:
            temp_excel.write(original_excel_bytes)
            temp_excel_path = temp_excel.name

        try:
            # Load Excel to read pricelist names from tab 2
            logger.info("Loading Excel to read pricelist names from tab 2")
            wb = openpyxl.load_workbook(temp_excel_path, data_only=True)

            if len(wb.sheetnames) < 2:
                logger.warning("Second tab not found in workbook, cannot generate CSVs")
                return []

            ws2 = wb.worksheets[1]  # Second tab (pricelists)

            # Read pricelist names from column A starting from row 3
            pricelist_names = []
            row = 3  # Start from row 3

            logger.info("Reading pricelist names from tab 2, column A starting from row 3")

            while row <= ws2.max_row:
                cell_value = ws2[f'A{row}'].value
                if cell_value is None or cell_value == '' or cell_value == 0:
                    break  # Stop when we hit empty cells

                pricelist_name = str(cell_value).strip()
                logger.info(f"Found pricelist at row {row}: '{pricelist_name}'")

                pricelist_names.append((pricelist_name, row))
                logger.info(f"Added pricelist: '{pricelist_name}' at row {row}")

                row += 1

                # Limit to 5 pricelists to prevent excessive processing
                if len(pricelist_names) >= 5:
                    logger.info("Reached limit of 5 pricelists")
                    break

            wb.close()

            logger.info(f"Found {len(pricelist_names)} pricelists to process")

            # Generate CSV for each pricelist using parallel processing
            from concurrent.futures import ProcessPoolExecutor

            csvs = []
            max_workers = min(len(pricelist_names), 5)  # Limit to 5 concurrent Excel processes

            logger.info(f"Starting parallel processing of {len(pricelist_names)} pricelists with {max_workers} workers")

            with ProcessPoolExecutor(max_workers=max_workers) as executor:
                # Submit all pricelist processing tasks and maintain order
                futures = [
                    executor.submit(process_pricelist_for_parallel, pricelist_name, pricelist_row, timestamp)
                    for pricelist_name, pricelist_row in pricelist_names
                ]

                logger.info(f"Submitted {len(futures)} pricelist Excel generation tasks to executor")

                # Collect Excel results from parallel processing
                excel_results = []
                completed_count = 0
                for future in futures:
                    try:
                        result = future.result()
                        completed_count += 1
                        pricelist_name, excel_bytes, result_timestamp = result
                        logger.info(f"Completed Excel generation for pricelist '{pricelist_name}' ({completed_count}/{len(futures)})")
                        excel_results.append((pricelist_name, excel_bytes, result_timestamp))
                    except Exception as exc:
                        # Get the pricelist name from the failed future (need to find which one it was)
                        pricelist_name = "unknown"
                        for i, f in enumerate(futures):
                            if f == future:
                                pricelist_name = pricelist_names[i][0]
                                break
                        logger.error(f'Pricelist {pricelist_name} generated an exception: {str(exc)}')
                        logger.error(f'Exception type: {type(exc).__name__}')
                        import traceback
                        logger.error(f'Exception traceback: {traceback.format_exc()}')
                        raise  # Re-raise to fail the entire operation

            logger.info(f"Parallel Excel generation completed - processed {completed_count} tasks")

            # Now convert Excel files to CSV in main process (xlwings can only work in main process)
            logger.info("Converting Excel files to CSV in main process...")
            for pricelist_name, excel_bytes, result_timestamp in excel_results:
                try:
                    logger.info(f"Converting Excel to CSV for pricelist '{pricelist_name}'")
                    csv_bytes = generate_csv_from_excel(excel_bytes)
                    csv_filename = f"justframeit_price_export_pricelist_{pricelist_name.replace(' ', '_').lower()}_{result_timestamp}.csv"
                    logger.info(f"Successfully converted Excel to CSV for pricelist '{pricelist_name}'")
                    csvs.append((pricelist_name, csv_bytes, csv_filename))
                except Exception as e:
                    logger.error(f"Error converting Excel to CSV for pricelist '{pricelist_name}': {str(e)}")
                    raise

            logger.info(f"CSV conversion completed - generated {len(csvs)} CSV files")
            return csvs

        finally:
            # Clean up temporary file
            try:
                os.remove(temp_excel_path)
                logger.info(f"Cleaned up temporary Excel file: {temp_excel_path}")
            except Exception as e:
                logger.warning(f"Could not remove temporary Excel file: {e}")

    except Exception as e:
        logger.error(f"Error generating CSVs from pricelists: {str(e)}")
        logger.error(f"Exception type: {type(e).__name__}")
        import traceback
        logger.error(f"Exception traceback: {traceback.format_exc()}")
        raise

@price_export_v2_bp.route('/generate-price-export-v2', methods=['POST'])
def generate_price_export():
    """
    Generate CSV price-export directly by computing prices in Python.
    This approach does NOT use Excel templates - all calculations are done in Python.

    This route:
    1. Fetches products with Surface/Circumference price computation from Odoo
    2. Fetches pricelists and duration rules from Odoo
    3. Computes prices for each product x dimension combination using Python
    4. Generates CSV files for each pricelist
    5. Saves CSV files to Odoo's x_configuration binary fields
    6. Returns success/failure status

    POST request with payload containing x_studio_is_run_locally flag:
    POST /generate-price-export
    Content-Type: application/json
    {
        "_action": "CUSTOM - Product Price Export Trigger(#969)",
        "_id": 1,
        "_model": "x_configuration",
        "id": 1,
        "x_studio_is_run_locally": false
    }
    """
    try:
        # Get the request payload
        payload = request.get_json() or {}

        # Set up log capture
        log_handler, log_capture_string = create_log_capture_handler()
        logger.addHandler(log_handler)

        logger.info("Starting price-export generation route (DIRECT CSV - No Excel)")
        logger.info(f"Payload received: {json.dumps(payload, indent=2)}")

        # Get Odoo connection
        logger.info("Connecting to Odoo")
        uid = get_uid()
        models = get_odoo_models()
        logger.info("Connected to Odoo successfully")

        # Generate timestamp for filenames
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Generate CSVs directly using Python computation (no Excel needed)
        logger.info("Generating CSV files directly using Python computation...")
        all_pricelist_csvs = generate_csv_direct(models, uid)
        if not all_pricelist_csvs:
            all_pricelist_csvs = []
        logger.info(f"Generated {len(all_pricelist_csvs)} CSV files directly")

        # Get counts for reporting
        # Fetch products count
        products = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'product.product', 'search_count',
            [[['x_studio_price_computation', 'in', ['Surface', 'Circumference']]]]
        )
        total_products = products

        # Fetch pricelists count
        all_pricelists = models.execute_kw(
            ODOO_DB, uid, ODOO_API_KEY,
            'product.pricelist', 'search_read',
            [],
            {'fields': ['name'], 'limit': 100}
        )
        total_pricelists = len([p for p in all_pricelists if p.get('name', '').lower() != 'default'])

        # Set up CSV variables
        if all_pricelist_csvs:
            csv_bytes, pricelist_name, csv_filename = all_pricelist_csvs[0]
            additional_csvs = all_pricelist_csvs
        else:
            csv_bytes = None
            csv_filename = None
            additional_csvs = []

        # Find the x_configuration record to update
        logger.info("Finding x_configuration record")
        config_ids = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'search', [[]])

        if not config_ids:
            raise Exception("No x_configuration record found")

        config_id = config_ids[0]
        logger.info(f"Found configuration record ID: {config_id}")

        # Prepare update values for CSV files
        update_vals = {}

        # Add CSV fields for each pricelist
        additional_csv_info = []
        for i, (csv_bytes_data, pl_name, csv_filename_data) in enumerate(additional_csvs):
            csv_field_num = i + 1  # Field 1, 2, 3, etc.
            csv_base64_data = base64.b64encode(csv_bytes_data).decode('ascii')

            # Store in configuration fields (up to 5 pricelists)
            if csv_field_num <= 5:
                update_vals[f'x_studio_price_list_{csv_field_num}_csv'] = csv_base64_data
                update_vals[f'x_studio_price_list_{csv_field_num}_csv_filename'] = csv_filename_data

            additional_csv_info.append({
                'pricelist_name': pl_name,
                'field': f'x_studio_price_list_{csv_field_num}_csv',
                'filename': csv_filename_data
            })

            logger.info(f"Added CSV for pricelist '{pl_name}' to field x_studio_price_list_{csv_field_num}_csv")

        if update_vals:
            logger.info(f"Saving {len(additional_csvs)} CSV files to Odoo")
            models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
                'x_configuration', 'write', [config_id, update_vals])
            logger.info("CSV files saved to Odoo successfully")

        # Post to chatter with summary, logs, and attachments
        logger.info("Posting to configuration chatter")

        # Create attachments for CSV files and logs
        attachment_ids = []

        # Create attachments for all CSV files
        for csv_bytes_data, pl_name, csv_filename_data in additional_csvs:
            csv_base64_data = base64.b64encode(csv_bytes_data).decode('ascii')
            csv_attachment_data = {
                'name': csv_filename_data,
                'type': 'binary',
                'datas': csv_base64_data,
                'res_model': 'x_configuration',
                'res_id': config_id,
                'mimetype': 'text/csv'
            }
            csv_attachment_id = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
                'ir.attachment', 'create', [csv_attachment_data])
            attachment_ids.append(csv_attachment_id)
            logger.info(f"Created CSV attachment for pricelist '{pl_name}' with ID: {csv_attachment_id}")

        # Get captured logs and create log attachment
        captured_logs = log_capture_string.getvalue()
        log_filename = f"price_export_logs_{timestamp}.txt"
        log_attachment_data = {
            'name': log_filename,
            'type': 'binary',
            'datas': base64.b64encode(captured_logs.encode('utf-8')).decode('ascii'),
            'res_model': 'x_configuration',
            'res_id': config_id,
            'mimetype': 'text/plain'
        }
        log_attachment_id = models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'ir.attachment', 'create', [log_attachment_data])
        attachment_ids.append(log_attachment_id)
        logger.info(f"Created log attachment with ID: {log_attachment_id}")

        # Prepare response
        csv_files_count = len(additional_csvs)
        message = f'{csv_files_count} CSV files generated directly (Python computation) and saved to Odoo successfully'

        response_data = {
            'message': message,
            'config_id': config_id,
            'csv_filename': csv_filename,
            'additional_csvs': additional_csv_info,
            'total_csv_files': csv_files_count,
            'products_processed': total_products,
            'pricelists_processed': total_pricelists,
            'method': 'direct_python_computation',
            'status': 'success'
        }

        logger.info(f"Price-export generation completed - Config ID: {config_id}, Products: {total_products}, Pricelists: {total_pricelists}, CSV files: {csv_files_count}")

        # Get captured logs and clean up handler
        log_contents = log_capture_string.getvalue()
        logger.removeHandler(log_handler)
        log_capture_string.close()

        # Log the route call to Odoo logging model
        log_route_call(models, uid, '/generate-price-export', payload, log_contents, response_data)

        # Create comprehensive HTML chatter message
        csv_list_items = []
        for csv_info in additional_csv_info:
            csv_list_items.append(f"<li>Pricelist '{csv_info['pricelist_name']}': {csv_info['filename']}</li>")

        # Get dimensions count
        dimensions = get_default_dimensions()
        dimensions_count = len(dimensions)

        chatter_message = f"""<p><strong>✅ Price-Export Generation Completed Successfully</strong></p>

<p><strong>📝 Processing Details:</strong></p>

<p><strong>Method:</strong> Direct Python Computation (No Excel Required)</p>

<p><strong>Data Retrieved:</strong></p>
<ul>
<li>Fetched {total_products} products from Odoo</li>
<li>Retrieved {total_pricelists} pricelists</li>
<li>Using {dimensions_count} dimension combinations</li>
</ul>

<p><strong>Price Computation:</strong></p>
<ul>
<li>Formula: (base_cost + labor_cost) × (1 + margin)</li>
<li>Base cost: Surface or Circumference × standard_price</li>
<li>Labor cost: service_duration × cost_per_hour / 3600</li>
</ul>

<p><strong>CSV Generation:</strong></p>
<ul>
<li>Created {csv_files_count} CSV files total</li>
</ul>

<p><strong>Generated CSV Files:</strong></p>
<ul>{''.join(csv_list_items)}</ul>

<p><strong>Final Return Data:</strong></p>
<ul>
<li>message: '{response_data['message']}'</li>
<li>config_id: {config_id}</li>
<li>total_csv_files: {csv_files_count}</li>
<li>products_processed: {total_products}</li>
<li>pricelists_processed: {total_pricelists}</li>
<li>method: direct_python_computation</li>
<li>status: 'success'</li>
</ul>"""

        # Post message to chatter
        chatter_data = {
            'body': chatter_message,
            'body_is_html': True,
            'message_type': 'comment',
            'subtype_xmlid': 'mail.mt_note',
            'attachment_ids': attachment_ids,
        }

        models.execute_kw(ODOO_DB, uid, ODOO_API_KEY,
            'x_configuration', 'message_post', [config_id], chatter_data)

        logger.info("Successfully posted to configuration chatter")

        return jsonify(response_data)

    except Exception as e:
        # Clean up log handler and get logs if available
        log_contents = ""
        try:
            if 'log_handler' in locals() and 'log_capture_string' in locals():
                log_contents = log_capture_string.getvalue()
                logger.removeHandler(log_handler)
                log_capture_string.close()
        except:
            pass

        logger.error(f"Error in generate-price-export route: {str(e)}")
        logger.error(f"Exception type: {type(e).__name__}")
        import traceback
        logger.error(f"Exception traceback: {traceback.format_exc()}")

        # Prepare error response data
        error_response = {'error': str(e), 'status': 'error'}

        # Log the error to Odoo logging model
        log_route_call(None, None, '/generate-price-export', payload, log_contents, error_response)

        return jsonify(error_response), 500
